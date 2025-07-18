import base64
import io
import os
import re
from datetime import datetime

from dotenv import load_dotenv
import numpy as np
import torch
from openai import OpenAI
from openpyxl import load_workbook
from PIL import Image

import folder_paths


load_dotenv()


def create_openai_client() -> OpenAI:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise EnvironmentError("OPENAI_API_KEY not found in environment variables.")
    return OpenAI(api_key=api_key)


class OpenAIImageBatchGenerator:
    mdoels = ["dall-e-3", "gpt-image-1"]
    aspect_ratios = ["1:1", "3:2 (landscape)", "2:3 (portrait)"]

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "model": (cls.mdoels,),
                "aspect_ratio": (cls.aspect_ratios,),
                "batch_size": (
                    "INT",
                    {
                        "default": 1,
                        "min": 1,
                        "max": 10,
                        "step": 1
                    }
                ),
                "style_indication": ("STRING", {"multiline": True, "tooltip": "Style instructions for the image generation."}),
                "prompt_string": ("STRING", {"multiline": True}),
                "multiline": ("BOOLEAN", {"default": False}),
            }
        }

    RETURN_TYPES = ("IMAGE",)
    RETURN_NAMES = ("images",)
    FUNCTION = "generate_images"
    CATEGORY = "🐅cesilk_nodes"

    def generate_images(self, model, aspect_ratio, batch_size, style_indication, prompt_string, multiline):
        client = create_openai_client()

        size_map = {
            "gpt-image-1": {
                "1:1": "1024x1024",
                "3:2 (landscape)": "1536x1024",
                "2:3 (portrait)": "1024x1536"
            },
            "dall-e-3": {
                "1:1": "1024x1024",
                "3:2 (landscape)": "1792x1024",
                "2:3 (portrait)": "1024x1792"
            }
        }

        if aspect_ratio not in size_map[model]:
            raise Exception(f"Invalid aspect ratio '{aspect_ratio}' for model '{model}'")
    
        n = 1 if model == "dall-e-3" else batch_size

        resolved_size = size_map[model][aspect_ratio]

        if multiline:
            prompts = [prompt_string.strip()]
        else:
            prompts = [p.strip() for p in prompt_string.strip().split("\n") if p.strip()]

        images = []

        for prompt in prompts:
            styled_prompt = build_styled_prompt(style_indication, prompt)
            response = client.images.generate(
                model=model,
                prompt=styled_prompt,
                n=n,
                size=resolved_size,
                response_format="b64_json"
            )

            print(f"Successfully generated image for prompt: {styled_prompt}")

            # Get the generated image
            for img in response.data:
                if not img.b64_json:
                    raise Exception("No image data returned from OpenAI API")

                image_base64 = img.b64_json
                image_bytes = base64.b64decode(image_base64)

                # Convert to PIL Image
                generated_image = Image.open(io.BytesIO(image_bytes))

                # Convert to numpy array
                image_np = np.array(generated_image).astype(np.float32) / 255.0
                image_np = np.expand_dims(image_np, axis=0)  # Add batch dimension

                # Convert to torch tensor
                image_tensor = torch.from_numpy(image_np)

                images.append(image_tensor)

        batched_images = torch.cat(images, dim=0)
        return (batched_images,)


def build_styled_prompt(style_indication: str, prompt: str):
    return \
        f"Please follow the style instructions to generate the image.\n\n" + \
        f"# Style Instructions\n{style_indication}\n\n" + \
        f"# Description of generated image\n{prompt}"


class OpenAIImageDescriptionToTextfile:
    def __init__(self):
        self.output_dir = folder_paths.get_output_directory()

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "images": ("IMAGE", ),
                "prompt": ("STRING", {"multiline": True}),
                "save_textfile": ("BOOLEAN", {"default": False}),
                "filename_prefix": ("STRING", {"default": "ComfyUI"}),
                "save_excel": ("BOOLEAN", {"default": False}),
                "excel_path": ("STRING", {"default": "", "multiline": False, "tooltip": "Path to the Excel file."}),
                "sheet_name": ("STRING", {"default": "", "multiline": False}),
                "column": ("STRING", {"default": "A", "multiline": False, "tooltip": "Excel cell column name. Example: Column A"}),
                "start_row_num": (
                    "INT",
                    {
                        "default": 2,
                        "min": 1,
                        "max": 1000,
                        "step": 1,
                        "tooltip": "Excel cell row number."
                    }
                ),
            }
        }

    RETURN_TYPES = ()
    FUNCTION = "images_description_to_textfile"
    CATEGORY = "🐅cesilk_nodes"
    OUTPUT_NODE = True

    def images_description_to_textfile(self, images, prompt, save_textfile, filename_prefix, 
                                       save_excel, excel_path, sheet_name, column, start_row_num):
        client = create_openai_client()

        filename_prefix = self.apply_date_format(filename_prefix.strip())
        full_path = os.path.join(self.output_dir, filename_prefix)

        messages = []

        for index, image in enumerate(images):
            i = 255. * image.cpu().numpy()
            img = Image.fromarray(np.clip(i, 0, 255).astype(np.uint8))

            # Convert image to base64
            buffered = io.BytesIO()
            img.save(buffered, format="JPEG")
            base64_image = base64.b64encode(buffered.getvalue()).decode("utf-8")


            response = client.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    # {"role": "system", "content": ""},
                    {"role": "user", "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                            }
                        },
                    ]},
                ]
            )
            msg = response.choices[0].message.content
            messages.append(msg)

            if save_textfile:
                full_file_path = f"{full_path}_{index:04}.txt"
                os.makedirs(os.path.dirname(full_file_path), exist_ok=True)

                with open(full_file_path, "w", encoding="utf-8") as f:
                    f.write(msg)
                    print(f"Saved description to {full_file_path}")

        if save_excel:
            if not excel_path:
                raise ValueError("Excel path must be provided when save_excel is True.")

            wb = load_workbook(excel_path, data_only=True)
            ws = wb[sheet_name]

            for i, msg in enumerate(messages):
                cell = f"{column}{start_row_num + i}"
                ws[cell] = msg

            wb.save(excel_path)

        return {}, {}

    def apply_date_format(self, text):
        if r"%date:" in text:
            ymd = datetime.now().strftime("%Y-%m-%d")
            text = re.sub(r"%date:yyyy-MM-dd%", ymd, text)
            ymds = datetime.now().strftime("%Y%m%d%H%M%S")
            text = re.sub(r"%date:yyMMddhhmmss%", ymds, text)
        return text


class OpenAIChat:
    mdoels = ["gpt-4o", "gpt-4.1"]

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "model": (cls.mdoels,),
                "system_prompt": ("STRING", {"multiline": True, "tooltip": "System prompt to set the context for the chat."}),
                "user_prompt": ("STRING", {"multiline": True, "tooltip": "Prompt to send to the OpenAI API."}),
            }
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("text",)
    FUNCTION = "chat"
    CATEGORY = "🐅cesilk_nodes"

    def chat(self, model, system_prompt, user_prompt):
        client = create_openai_client()

        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ]
        )
        message = response.choices[0].message.content

        return (message,)
